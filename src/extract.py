import openpyxl, json, re, sys, os, unicodedata
from pathlib import Path

def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def sa(s):
    """strip accents + lowercase + strip whitespace, for label matching"""
    if not isinstance(s, str):
        return None
    return strip_accents(s).lower().strip()

def norm_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        return None  # formula errors like '#DIV/0!' etc.
    if isinstance(v, (int, float)):
        return round(v, 2) if isinstance(v, float) else v
    return v

def negate_kv(kv):
    if kv is None:
        return None
    out = {}
    for k in ("real", "budget", "prior_year"):
        out[k] = -kv[k] if isinstance(kv.get(k), (int, float)) else kv.get(k)
    out["delta_budget"] = (out["real"] - out["budget"]) if isinstance(out["real"], (int, float)) and isinstance(out["budget"], (int, float)) else None
    out["delta_prior_year"] = (out["real"] - out["prior_year"]) if isinstance(out["real"], (int, float)) and isinstance(out["prior_year"], (int, float)) else None
    return out

# Fixed 5-dealer roster for Groupeautomax. Order here is the permanent
# column/row/color order in the dashboard, regardless of upload order.
CANONICAL_DEALERS = [
    ("bmw", "BMW Sherbrooke", lambda s: "bmw" in s),
    ("stm", "STM (Ste-Marie Auto)", lambda s: "ste-marie" in s or "ste marie" in s or re.search(r'\bstm\b', s)),
    ("hawks", "HAWKS", lambda s: "hawk" in s),
    ("vw", "Volkswagen", lambda s: "volkswagen" in s or "volks" in s or re.search(r'\bvw\b', s)),
    ("hyundai", "Hyundai", lambda s: "hyundai" in s),
]

def canonicalize_dealer(company_name):
    s = strip_accents(company_name or '').lower()
    for key, display, match_fn in CANONICAL_DEALERS:
        if match_fn(s):
            return key, display
    return company_name, company_name

# ---------------------------------------------------------------------------
# Universal parser for the "Réalisé Mois" / "Print_Mois" (and AAD equivalents)
# sheets. Every Groupeautomax dealer file shares this exact underlying
# corporate template (only the sheet's name differs: BMW/STM call it
# "Réalisé Mois"/"Réalisé AAD", others call it "Print_Mois"/"Print_AAD"), so
# one parser covers all 5 dealers -- including the department-level detail
# (Véhicules neufs, Véhicules usagés, Service, Carrosserie, Pièces, Boutique)
# that the simpler "Résumé" sheet (BMW/STM only) never exposed.
# ---------------------------------------------------------------------------

MONTH_SHEET_CANDIDATES = ["Réalisé Mois", "Print_Mois"]
YTD_SHEET_CANDIDATES = ["Réalisé AAD", "Print_AAD"]

# department block: budget/real/delta/prior/delta columns (F,L,N,Q,S)
DEPT_COLS = {"budget": 6, "real": 12, "delta_budget": 14, "prior_year": 17, "delta_prior_year": 19}
# department block units columns (D,J,M,O,R)
UNIT_COLS = {"budget": 4, "real": 10, "delta_budget": 13, "prior_year": 15, "delta_prior_year": 18}
# company-wide summary block columns (Z,AF,AH,AL,AN)
SUMMARY_COLS = {"budget": 26, "real": 32, "delta_budget": 34, "prior_year": 38, "delta_prior_year": 40}
# department expense-total rows also carry a "% of department profit brut"
# figure one column to the left of each $ column (E budget%, K réel%, P prior%)
DEPT_PCT_COLS = {"budget": 5, "real": 11, "prior_year": 16}

DEPT_ROW_LABELS = {
    "total_variables": "total variables",
    "total_personnel": "total personnel",
    "total_semifixes": "total semi-fixes",
    "total_depenses": "total des depenses",
    "autres_revenus": "autres revenus",
    "profit_departemental": "profit departemental",
}
# Which of the rows above carry the "% of profit brut" figure (the revenue
# lines -- autres_revenus, profit_departemental -- don't have this column).
DEPT_PCT_KEYS = {"total_variables", "total_personnel", "total_semifixes", "total_depenses"}
# Which line represents the department's true "units sold" figure differs by
# department, because of what the department total (gross_row) does or
# doesn't fold in:
#  - Véhicules neufs: "Total Dépt. Neufs" = Total Neufs (retail) + Ex-Démos
#    et courtoisies + Flottes, with NO wholesale line in this department --
#    so the department total itself is exactly the right unit count (verified
#    against BMW's June AAD: 163 = 118 retail + 28 démo + 17 flottes, matching
#    what BMW's own reporting calls "ventes de neufs").
#  - Véhicules usagés: "Total Dépt. Usagés" ADDS a separate "Ventes au Gros"
#    (wholesale) line on top of retail -- wholesale is a different sales
#    channel (dealer-to-dealer, not a customer sale), so folding it in
#    inflated the count (52 instead of the validated 28 for BMW). Usagés
#    keeps reading the named retail-only line instead of the department total.
GROSS_ROW_UNITS_DEPTS = {"Véhicules neufs"}
UNITS_DEPT_LINES = {
    "Véhicules usagés": "total usages",
}
# Extra standalone unit line, when present, surfaced as its own KPI
# ("Unités flottes") rather than folded into the retail count above.
FLEET_LINE = "flottes"

SUMMARY_ROW_LABELS = {
    "ventes_nettes": "ventes nettes",
    "profit_brut": "profit brut",
    "total_depenses": "total des depenses",
    "total_autres_revenus": "total autres revenus",
    "baiia_operationnel": "baiia operationnel",
    "profit_net": "profit net",
    "impot": "impot exigible",
    "profit_net_apres_impot": "profit net apres impot",
}

def ratio_kv(numerator, denominator):
    """numerator/denominator as a 0..1 ratio for real/budget/prior_year, with
    point-deltas -- used to derive EBT % of gross profit / % of sales (ROS)
    since the company summary block doesn't carry these as their own sheet
    columns the way department expense-total rows do."""
    if not numerator or not denominator:
        return None
    out = {}
    for k in ("real", "budget", "prior_year"):
        n, d = numerator.get(k), denominator.get(k)
        out[k] = (n / d) if isinstance(n, (int, float)) and isinstance(d, (int, float)) and d else None
    out["delta_budget"] = (out["real"] - out["budget"]) if isinstance(out["real"], (int, float)) and isinstance(out["budget"], (int, float)) else None
    out["delta_prior_year"] = (out["real"] - out["prior_year"]) if isinstance(out["real"], (int, float)) and isinstance(out["prior_year"], (int, float)) else None
    return out

def read_kv(ws, row, colmap):
    if row is None:
        return {k: None for k in ("real", "budget", "delta_budget", "prior_year", "delta_prior_year")}
    return {k: norm_num(ws.cell(row=row, column=c).value) for k, c in colmap.items()}

def read_pct(ws, row):
    """Read the '% of department profit brut' figure that rides alongside an
    expense-total row (budget/real/prior, as a 0..1 ratio), with point-deltas
    computed rather than trusting any adjacent sheet formula."""
    vals = {k: norm_num(ws.cell(row=row, column=c).value) for k, c in DEPT_PCT_COLS.items()}
    real, budget, prior = vals.get("real"), vals.get("budget"), vals.get("prior_year")
    vals["delta_budget"] = (real - budget) if isinstance(real, (int, float)) and isinstance(budget, (int, float)) else None
    vals["delta_prior_year"] = (real - prior) if isinstance(real, (int, float)) and isinstance(prior, (int, float)) else None
    return vals

def find_first_sheet(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None

def find_department_blocks(ws, max_row=400):
    """Return {dept_name: [(start_row, end_row), ...]} by scanning column B
    for named runs. Fragments of the same name are kept as SEPARATE ranges
    rather than collapsed into one min..max span: collapsing is wrong when a
    stray single-row fragment of one department sits physically inside
    another department's block (seen in the wild -- BMW's "Rachat bail" row
    is tagged "Véhicules neufs" in column B but is physically printed in the
    middle of the Véhicules usagés rows). Collapsing to one span would
    silently swallow the other department's rows into this one's range;
    keeping fragments separate means each department's own row-scans only
    ever visit rows actually tagged with its name."""
    starts = []
    last_name = None
    for r in range(1, max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip() and b != last_name:
            starts.append((b, r))
            last_name = b
    bounds = {}
    for i, (name, s) in enumerate(starts):
        e = (starts[i + 1][1] - 1) if i + 1 < len(starts) else max_row
        bounds.setdefault(name, []).append((s, e))
    return bounds

def dept_rows(fragments):
    for s, e in fragments:
        for r in range(s, e + 1):
            yield r

def is_total_label(label):
    s = sa(label)
    return s.startswith("total") or s in ("profit departemental", "autres revenus")

def extract_line_items(ws, rows, start, end, gross_row, depenses_row, label_rows):
    """Every individually labeled row belonging to this department, verbatim --
    e.g. Autos détail / Camions détail / Total Neufs / Ex-Démos et
    courtoisies / Flottes -- so the dashboard can show the full source
    breakdown behind each rolled-up metric, not just the rollup itself.
    Each item is tagged with the section it belongs to (matching one of the
    top-line DEPARTMENT_METRICS keys) so the UI can show "the detail behind
    this specific card" rather than one undifferentiated wall of rows.
    Iterates `rows` (this department's own fragment rows) rather than a
    blind range(start, end) -- a disjoint department can sit in the gap
    between two of this department's fragments."""
    boundaries = [("ventes", start, gross_row or end)]
    if depenses_row:
        seg_start = depenses_row
        for name, key in (("variables", "total_variables"), ("personnel", "total_personnel"), ("semifixes", "total_semifixes")):
            row = label_rows.get(key)
            if row:
                boundaries.append((name, seg_start, row))
                seg_start = row + 1
        td = label_rows.get("total_depenses")
        if td:
            boundaries.append(("depenses_total", seg_start, td))
            boundaries.append(("autres", td + 1, end))
        else:
            boundaries.append(("autres", seg_start, end))

    def section_for(r):
        for name, s, e in boundaries:
            if s <= r <= e:
                return name
        return "autres"

    items = []
    for r in rows:
        label = ws.cell(row=r, column=3).value
        if not isinstance(label, str) or not label.strip():
            continue
        if label.strip() == "[...]":
            continue  # unused template placeholder slot, not a real line
        units = read_kv(ws, r, UNIT_COLS)
        money = read_kv(ws, r, DEPT_COLS)
        pct = read_pct(ws, r)
        has_units = any(v is not None for v in units.values())
        has_money = any(v is not None for v in money.values())
        has_pct = any(v is not None for v in pct.values())
        if not (has_units or has_money or has_pct):
            continue  # pure section header (e.g. "DEPENSES"), no data of its own
        items.append({
            "label": label.strip(),
            "section": section_for(r),
            "is_total": is_total_label(label),
            "units": units if has_units else None,
            "money": money if has_money else None,
            "pct": pct if has_pct else None,
        })
    return items

def analyze_department(ws, name, fragments):
    rows = list(dept_rows(fragments))
    start, end = fragments[0][0], fragments[-1][1]

    depenses_row = None
    for r in rows:
        if sa(ws.cell(row=r, column=3).value) == "depenses":
            depenses_row = r
            break

    gross_row = None
    if depenses_row:
        # search backward for the nearest labeled $ row, but never past the
        # start of the fragment depenses_row itself belongs to -- crossing
        # into a different (earlier) fragment could cross into a gap
        # occupied by another department's rows.
        frag_start = next(s for s, e in fragments if s <= depenses_row <= e)
        for r in range(depenses_row - 1, frag_start - 1, -1):
            c = ws.cell(row=r, column=3).value
            f = ws.cell(row=r, column=6).value
            if isinstance(c, str) and c.strip() and f is not None:
                gross_row = r
                break

    label_rows = {}
    for r in rows:
        key_label = sa(ws.cell(row=r, column=3).value)
        for key, target in DEPT_ROW_LABELS.items():
            if key_label == target and key not in label_rows:
                label_rows[key] = r

    result = {}
    if gross_row:
        result["profit_brut"] = read_kv(ws, gross_row, DEPT_COLS)

    units_target = UNITS_DEPT_LINES.get(name)
    fleet_row = None
    scan_rows = [r for r in rows if r <= (depenses_row or end)]
    if name in GROSS_ROW_UNITS_DEPTS and gross_row:
        result["units"] = read_kv(ws, gross_row, UNIT_COLS)
        for r in scan_rows:
            if sa(ws.cell(row=r, column=3).value) == FLEET_LINE:
                fleet_row = r
                break
    elif units_target:
        for r in scan_rows:
            label = sa(ws.cell(row=r, column=3).value)
            if label == units_target:
                result["units"] = read_kv(ws, r, UNIT_COLS)
            elif label == FLEET_LINE:
                fleet_row = r
    if fleet_row:
        result["units_flottes"] = read_kv(ws, fleet_row, UNIT_COLS)

    for key, row in label_rows.items():
        result[key] = read_kv(ws, row, DEPT_COLS)
        if key in DEPT_PCT_KEYS:
            result[key]["pct"] = read_pct(ws, row)

    result["line_items"] = extract_line_items(ws, rows, start, end, gross_row, depenses_row, label_rows)
    return result

def extract_departments(ws):
    if ws is None:
        return {}
    bounds = find_department_blocks(ws)
    departments = {}
    for name, fragments in bounds.items():
        data = analyze_department(ws, name, fragments)
        if data:
            departments[name] = data
    return departments

def extract_summary(ws):
    if ws is None:
        return {}
    rows = {}
    for r in range(1, ws.max_row + 1):
        label = sa(ws.cell(row=r, column=22).value)  # column V
        if label is None:
            continue
        for key, target in SUMMARY_ROW_LABELS.items():
            if label == target and key not in rows:
                rows[key] = r
    return {key: read_kv(ws, row, SUMMARY_COLS) for key, row in rows.items()}

# ---------------------------------------------------------------------------
# Legacy "Résumé" sheet parser (BMW / STM only) -- kept solely to harvest the
# quarter ("Q2", "Q3", ...) blocks it contains, since the universal template
# above has no quarterly rollup of its own.
# ---------------------------------------------------------------------------

def classify_resume_section(title):
    t = sa(title)
    if "annee a date" in t or "aad" in t:
        return "ytd"
    m = re.match(r'^q(\d)$', t.strip())
    if m:
        return f"quarter_q{m.group(1)}"
    return "month"

RESUME_CANON_RULES = [
    (lambda s: 'unit' in s and 'neuf' in s, 'unites_neuf', 'Unités neuf', 'volume'),
    (lambda s: 'unit' in s and 'usag' in s, 'unites_usage', 'Unités usagé', 'volume'),
    (lambda s: 'unit' in s and 'flotte' in s, 'unites_flottes', 'Unités flottes', 'volume'),
    (lambda s: 'total profit brut' in s, 'pb_total', 'Total profit brut', 'profit'),
    (lambda s: 'profit brut' in s and 'neuf' in s, 'pb_neuf', 'Profit brut neuf', 'profit'),
    (lambda s: 'profit brut' in s and 'usag' in s, 'pb_usage', 'Profit brut usagé', 'profit'),
    (lambda s: 'profit brut' in s and 'wholesale' in s, 'pb_wholesale', 'Profit brut wholesale autres', 'profit'),
    (lambda s: 'profit brut' in s and 'service' in s, 'pb_service', 'Profit brut service', 'profit'),
    (lambda s: 'profit brut' in s and 'saaq' in s, 'pb_saaq', 'Profit brut SAAQ', 'profit'),
    (lambda s: 'profit brut' in s and 'carross' in s, 'pb_carrosserie', 'Profit brut carrosserie', 'profit'),
    (lambda s: 'profit brut' in s and ('piece' in s or 'pièce' in s), 'pb_pieces', 'Profit brut pièces', 'profit'),
    (lambda s: 'autres revenus' in s, 'autres_revenus', 'Autres revenus', 'revenue_expense'),
    (lambda s: s.strip() == 'depenses' or s.strip().startswith('depenses'), 'depenses', 'Dépenses', 'revenue_expense'),
    (lambda s: 'amort' in s, 'amort_interets', 'Amortissement et intérêts', 'revenue_expense'),
    (lambda s: 'profit net' in s, 'profit_net', 'Profit net', 'net'),
]

def canonicalize_resume_label(label):
    s = sa(label)
    for match_fn, key, disp, group in RESUME_CANON_RULES:
        if match_fn(s):
            return key, disp, group
    return None, label, 'other'

def extract_resume_quarters(wb):
    """Return {'quarter_qN': {'source_title':..., 'kpis': {...}}} if a
    Résumé sheet with quarter blocks exists; {} otherwise."""
    if 'Résumé' not in wb.sheetnames:
        return {}
    ws = wb['Résumé']
    sections = {}
    current_section = None
    r = 1
    while r <= ws.max_row:
        row = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        colB, colC, colD, colE = row[1], row[2], row[3], row[4]
        if colB is None and isinstance(colC, str) and colD is None and colE is None:
            current_section = colC.strip()
            sections[current_section] = {"type": classify_resume_section(current_section), "kpis": {}}
            r += 1
            continue
        if current_section and not sections[current_section]["kpis"] and colB is None and isinstance(colC, str) \
           and 'reel' in sa(colC):
            r += 1
            continue
        if current_section and isinstance(colB, str) and colB.strip():
            label = colB.strip()
            key, disp, group = canonicalize_resume_label(label)
            entry = {
                "label": disp, "raw_label": label, "group": group,
                "real": norm_num(row[2]), "budget": norm_num(row[3]), "delta_budget": norm_num(row[4]),
                "prior_year": norm_num(row[5]), "delta_prior_year": norm_num(row[6]),
            }
            sections[current_section]["kpis"][key or f"other::{label}"] = entry
            r += 1
            continue
        r += 1
    return {
        sec["type"]: {"source_title": title, "kpis": sec["kpis"]}
        for title, sec in sections.items() if sec["type"].startswith("quarter_")
    }

# ---------------------------------------------------------------------------
# Top-level extraction
# ---------------------------------------------------------------------------

def department_kpis_for_company_view(departments):
    """Map the rich per-department detail onto the flat company-wide KPI set
    used by the existing Volume / Profit brut / Revenus & dépenses / Profit
    net tabs, so all 5 dealers populate those tabs consistently."""
    kpis = {}
    dept_to_canon = {
        "Véhicules neufs": ("pb_neuf", "Profit brut neuf"),
        "Véhicules usagés": ("pb_usage", "Profit brut usagé"),
        "Service": ("pb_service", "Profit brut service"),
        "Carrosserie": ("pb_carrosserie", "Profit brut carrosserie"),
        "Pièces": ("pb_pieces", "Profit brut pièces"),
    }
    for dept_name, (key, label) in dept_to_canon.items():
        dept = departments.get(dept_name)
        if dept and dept.get("profit_brut"):
            kpis[key] = dict(dept["profit_brut"], label=label, raw_label=dept_name, group="profit")

    units_map = {
        "Véhicules neufs": ("unites_neuf", "Unités neuf"),
        "Véhicules usagés": ("unites_usage", "Unités usagé"),
    }
    for dept_name, (key, label) in units_map.items():
        dept = departments.get(dept_name)
        if dept and dept.get("units"):
            u = dept["units"]
            kpis[key] = {
                "label": label, "raw_label": dept_name, "group": "volume",
                "real": u["real"], "budget": u["budget"], "delta_budget": u["delta_budget"],
                "prior_year": u["prior_year"], "delta_prior_year": u["delta_prior_year"],
            }

    neuf = departments.get("Véhicules neufs")
    if neuf and neuf.get("units_flottes"):
        u = neuf["units_flottes"]
        kpis["unites_flottes"] = {
            "label": "Unités flottes", "raw_label": "Flottes", "group": "volume",
            "real": u["real"], "budget": u["budget"], "delta_budget": u["delta_budget"],
            "prior_year": u["prior_year"], "delta_prior_year": u["delta_prior_year"],
        }
    return kpis

def summary_kpis_for_company_view(summary):
    kpis = {}
    if summary.get("profit_brut"):
        kpis["pb_total"] = dict(summary["profit_brut"], label="Total profit brut", raw_label="Profit brut", group="profit")
    if summary.get("total_autres_revenus"):
        kpis["autres_revenus"] = dict(summary["total_autres_revenus"], label="Autres revenus", raw_label="Total autres revenus", group="revenue_expense")
    if summary.get("total_depenses"):
        kpis["depenses"] = dict(negate_kv(summary["total_depenses"]), label="Dépenses", raw_label="Total des dépenses", group="revenue_expense")
    if summary.get("profit_net"):
        kpis["profit_net"] = dict(summary["profit_net"], label="Profit net", raw_label="Profit Net", group="net")

    # --- EBITDA / EBT section: company-wide profitability rollup, mirroring
    # what Quotus calls "Earnings before tax" (EBITDAR $, EBT $, EBT % of
    # Gross Profit, EBT % of Sales / ROS). "Profit Net" in this template is
    # BEFORE income tax (it sits above "Impôt exigible" in the sheet), so it
    # IS the EBT figure -- exposed a second time under its own key so it can
    # live in this dedicated tab without disturbing the existing "Profit net"
    # KPI under the net tab.
    if summary.get("ventes_nettes"):
        kpis["ventes_nettes"] = dict(summary["ventes_nettes"], label="Ventes nettes", raw_label="Ventes nettes", group="ebitda")
    if summary.get("baiia_operationnel"):
        kpis["ebitda"] = dict(summary["baiia_operationnel"], label="BAIIA (EBITDA)", raw_label="BAIIA Opérationnel", group="ebitda")
    if summary.get("profit_net"):
        kpis["ebt"] = dict(summary["profit_net"], label="Profit net avant impôt (EBT)", raw_label="Profit Net", group="ebitda")
    if summary.get("impot"):
        kpis["impot"] = dict(summary["impot"], label="Impôt exigible", raw_label="Impôt exigible", group="ebitda")
    if summary.get("profit_net_apres_impot"):
        kpis["profit_net_apres_impot"] = dict(summary["profit_net_apres_impot"], label="Profit net après impôt", raw_label="Profit Net après impôt", group="ebitda")
    ebt_pct_gp = ratio_kv(summary.get("profit_net"), summary.get("profit_brut"))
    if ebt_pct_gp:
        kpis["ebt_pct_profit_brut"] = dict(ebt_pct_gp, label="EBT % du profit brut", raw_label="EBT % of Gross Profit", group="ebitda", format="percent")
    ros = ratio_kv(summary.get("profit_net"), summary.get("ventes_nettes"))
    if ros:
        kpis["ros"] = dict(ros, label="EBT % des ventes (ROS)", raw_label="EBT % of Sales (ROS)", group="ebitda", format="percent")

    return kpis

def sum_kv(kv_list):
    """Sum a list of kv dicts (real/budget/prior_year) field-by-field,
    ignoring any dict where a given field is missing/non-numeric -- mirrors
    the dashboard's own client-side sumKv() so server-extracted combined
    KPIs and client-combined dealer totals agree."""
    kv_list = [kv for kv in kv_list if kv]
    if not kv_list:
        return None
    out = {}
    for k in ("real", "budget", "prior_year"):
        vals = [kv.get(k) for kv in kv_list if isinstance(kv.get(k), (int, float))]
        out[k] = sum(vals) if vals else None
    out["delta_budget"] = (out["real"] - out["budget"]) if isinstance(out["real"], (int, float)) and isinstance(out["budget"], (int, float)) else None
    out["delta_prior_year"] = (out["real"] - out["prior_year"]) if isinstance(out["real"], (int, float)) and isinstance(out["prior_year"], (int, float)) else None
    return out

def expense_breakdown_kpis_for_company_view(departments):
    """Break the company-wide "Dépenses" total (revenue_expense tab) down by
    category -- variables / personnel / semi-fixes -- by summing each
    department's own subtotal (already extracted per-department, just never
    rolled up to the company level before). Negated like "Dépenses" itself so
    all three read as reductions alongside it, not raw expense magnitudes."""
    kpis = {}
    specs = (
        ("total_variables", "depenses_variables", "Dépenses variables"),
        ("total_personnel", "depenses_personnel", "Dépenses de personnel"),
        ("total_semifixes", "depenses_semifixes", "Dépenses semi-fixes"),
    )
    for dept_key, kpi_key, label in specs:
        combined = sum_kv([dept.get(dept_key) for dept in departments.values()])
        if combined:
            kpis[kpi_key] = dict(negate_kv(combined), label=label, raw_label=label, group="revenue_expense")
    return kpis

def unit_economics_kpis_for_company_view(kpis):
    """Profit brut par unité (GPA/PVR) for the two departments where a
    per-unit figure is meaningful -- derived from figures already extracted:
    profit brut ÷ unités."""
    out = {}
    specs = (
        ("pb_neuf", "unites_neuf", "gpa_neuf", "Profit brut par unité (neuf)"),
        ("pb_usage", "unites_usage", "gpa_usage", "Profit brut par unité (usagé)"),
    )
    for pb_key, units_key, kpi_key, label in specs:
        ratio = ratio_kv(kpis.get(pb_key), kpis.get(units_key))
        if ratio:
            out[kpi_key] = dict(ratio, label=label, raw_label=label, group="profit")
    return out

def extract_period(wb, sheet_candidates):
    ws = find_first_sheet(wb, sheet_candidates)
    if ws is None:
        return None
    departments = extract_departments(ws)
    summary = extract_summary(ws)
    kpis = {}
    kpis.update(department_kpis_for_company_view(departments))
    kpis.update(summary_kpis_for_company_view(summary))
    kpis.update(expense_breakdown_kpis_for_company_view(departments))
    kpis.update(unit_economics_kpis_for_company_view(kpis))
    return {"kpis": kpis, "departments": departments, "summary": summary}

def extract_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    p = wb['Parametres']
    company = p['B1'].value
    month_name = p['B3'].value
    month_num = p['C3'].value
    year = p['B4'].value

    sections = {}
    month_data = extract_period(wb, MONTH_SHEET_CANDIDATES)
    if month_data:
        sections["month"] = {"source_title": month_name, "kpis": month_data["kpis"], "departments": month_data["departments"]}
    ytd_data = extract_period(wb, YTD_SHEET_CANDIDATES)
    if ytd_data:
        sections["ytd"] = {"source_title": f"AAD {month_name}", "kpis": ytd_data["kpis"], "departments": ytd_data["departments"]}

    sections.update(extract_resume_quarters(wb))

    return {
        "company": company.strip() if isinstance(company, str) else company,
        "month_name": month_name,
        "month_num": month_num,
        "year": year,
        "period_key": f"{year}-{int(month_num):02d}",
        "sections": sections,
    }

# ---------------------------------------------------------------------------
# HAWKS-specific parser.
#
# HAWKS' monthly file is GM Canada's standardized "Composite Financial
# Statement" (GMCL) export -- a completely different workbook from the
# Quotus-based template every other Groupeautomax dealer sends (different
# sheet names: Page1..Page8, Page5_<brand>_<MCI|VE>, etc; different row/
# column layout; and, critically, NO budget or prior-year columns at all --
# only "Mois" (this period's actual) and "Cumul annuel" (year-to-date
# actual)). So HAWKS needs its own reader here, but it still funnels into
# the exact same department_kpis_for_company_view() / summary_kpis_for_
# company_view() aggregation the other 5 dealers use, by building
# `departments` / `summary` dicts shaped exactly like theirs (with budget
# and prior_year simply always None for HAWKS -- shown as "n/d" on the
# dashboard, same as any other missing comparison).
# ---------------------------------------------------------------------------

HAWKS_MONTHS = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

# The 8 per-brand "new vehicle" detail sheets whose "TOTAL VÉHICULES NEUFS"
# row must be summed for the dealership-wide new-unit count (each brand's
# total already folds in that brand's fleet/parcs units -- see
# hawks_fleet_units below, which reads the *same* units as a separate memo
# breakout, not an additional amount to add on top).
HAWKS_NEW_VEHICLE_BRAND_SHEETS = [
    "Page5_CHV_MCI", "Page5_CHV_VE", "Page5_BUI_MCI", "Page5_BUI_VE",
    "Page5_GMC_MCI", "Page5_GMC_VE", "Page5_CAD_MCI", "Page5_CAD_VE",
]

# Fixed row numbers within Page3 ("Véhicules neufs" / "Véhicules d'occasion")
# and Page4 ("Mécanique" / "Carrosserie" / "Pièces et accessoires") -- these
# are printed-form line numbers on GM's standardized statement, stable across
# reporting periods for a given dealer (verified against the December 2025,
# June 2026 and July 2026 HAWKS files, both sheets, every department/column).
#
# NOTE on "total_semifixes": row 59 ("TOTAL FRAIS GÉNÉRAUX FIXES") is NOT
# semi-fixes+fixe alone -- it's personnel + semi-fixes + fixe combined
# (verified: row20 + row43 + row58 == row59 to the dollar, every file/dept/
# column checked). Reading it directly here would double-count personnel,
# which is already reported separately under total_personnel. The
# Quotus-comparable "semi-fixes" figure (excluding personnel) is built from
# rows 43 + 58 alone -- see hawks_department_section() below, which
# overrides this row-59 default after the dict comprehension runs.
HAWKS_DEPT_ROWS = {
    "profit_brut": 5,
    "total_variables": 10,
    "total_personnel": 20,
    "total_semifixes": 59,  # placeholder -- overridden below with rows 43+58
    "total_depenses": 60,
    "profit_departemental": 64,
}

# Same fixed row layout as HAWKS_DEPT_ROWS above, broken down into the
# individually labeled source lines behind each subtotal -- lets the
# dashboard's "Voir le détail des postes" drill-down work for HAWKS the same
# way it already does for the Quotus-template dealers. Row ranges are
# inclusive of their closing subtotal row (e.g. "variables" ends at row 10,
# TOTAL FRAIS VARIABLES itself), matching extract_line_items()'s convention
# for the Quotus parser. Rows 44-58 (GM's "fixe" block) are tagged
# "semifixes" too, not a separate section -- since total_semifixes is now
# built from rows 43+58 combined (see note above), the line items need to
# sum to that same figure. Rows 59/60 (both redundant combined subtotals)
# are deliberately excluded -- nothing above needs a stand-alone line for
# them.
HAWKS_LINE_ITEM_SECTIONS = (
    ("ventes", range(4, 6)),
    ("variables", range(7, 11)),
    ("personnel", range(11, 21)),
    ("semifixes", range(21, 59)),
    ("autres", range(61, 65)),
)
HAWKS_LINE_ITEM_TOTAL_ROWS = {10, 20, 43, 58, 64}

def real_only_kv(value):
    """A real-only kv (no budget/prior-year exists in HAWKS' source file)."""
    v = norm_num(value)
    return {"real": v, "budget": None, "delta_budget": None, "prior_year": None, "delta_prior_year": None}

def hawks_find_row(ws, label_prefix, col=3, max_row=200):
    target = sa(label_prefix)
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and sa(v).startswith(target):
            return r
    return None

def hawks_new_vehicle_units(wb):
    """Sum "TOTAL VÉHICULES NEUFS" across the 8 per-brand sheets -> (month_kv, ytd_kv)."""
    month_total, ytd_total, any_found = 0, 0, False
    for name in HAWKS_NEW_VEHICLE_BRAND_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        row = hawks_find_row(ws, "total vehicules neufs")
        if row is None:
            continue
        any_found = True
        month_total += ws.cell(row=row, column=6).value or 0
        ytd_total += ws.cell(row=row, column=11).value or 0
    if not any_found:
        return None, None
    return real_only_kv(month_total), real_only_kv(ytd_total)

def hawks_fleet_units(wb):
    """"Unités flottes" memo -- TOTAL VÉHICULES DE PARCS ET GOUV, already
    included inside hawks_new_vehicle_units' brand totals above, exposed here
    a second time as its own KPI (mirrors how the Quotus template surfaces
    "Flottes" alongside, not subtracted from, the neuf total)."""
    if "Page5_PARCS_GOUV" not in wb.sheetnames:
        return None, None
    ws = wb["Page5_PARCS_GOUV"]
    row = hawks_find_row(ws, "total vehicules de parcs et gouv")
    if row is None:
        return None, None
    return real_only_kv(ws.cell(row=row, column=6).value), real_only_kv(ws.cell(row=row, column=11).value)

def hawks_used_vehicle_units(wb):
    """TOTAL VÉH. D'OCCASION DÉTAIL (retail only, excludes wholesale "en
    gros" -- same convention as the Quotus template's unites_usage)."""
    ws = wb["Page6"]
    row = hawks_find_row(ws, "total veh. d'occasion detail") or hawks_find_row(ws, "total veh")
    if row is None:
        return None, None
    return real_only_kv(ws.cell(row=row, column=6).value), real_only_kv(ws.cell(row=row, column=11).value)

def hawks_line_items(ws, money_col):
    """Every individually labeled row behind a HAWKS department's subtotals,
    tagged by section exactly like extract_line_items() does for the Quotus
    template -- see HAWKS_LINE_ITEM_SECTIONS above for the row->section map
    and why rows 44-58 fold into "semifixes" rather than a section of their
    own."""
    items = []
    for section, rows in HAWKS_LINE_ITEM_SECTIONS:
        for r in rows:
            label = ws.cell(row=r, column=2).value
            if not isinstance(label, str) or not label.strip():
                continue
            money = real_only_kv(ws.cell(row=r, column=money_col).value)
            if money["real"] is None:
                continue
            items.append({
                "label": label.strip(),
                "section": section,
                "is_total": r in HAWKS_LINE_ITEM_TOTAL_ROWS,
                "units": None,
                "money": money,
                "pct": None,
            })
    return items

def hawks_department_section(wb, sheet_name, money_col, units_kv=None, units_flottes_kv=None):
    """One department's data for one section (month or ytd), keyed exactly
    like analyze_department()'s output so it plugs into
    department_kpis_for_company_view() unchanged."""
    ws = wb[sheet_name]
    data = {key: real_only_kv(ws.cell(row=row, column=money_col).value) for key, row in HAWKS_DEPT_ROWS.items()}

    # Override the row-59 placeholder with the Quotus-comparable semi-fixes
    # figure (rows 43 + 58, excluding personnel -- see HAWKS_DEPT_ROWS note).
    semifixes_true = ws.cell(row=43, column=money_col).value
    fixe_true = ws.cell(row=58, column=money_col).value
    if isinstance(semifixes_true, (int, float)) or isinstance(fixe_true, (int, float)):
        data["total_semifixes"] = real_only_kv((semifixes_true or 0) + (fixe_true or 0))

    if units_kv is not None:
        data["units"] = units_kv
    if units_flottes_kv is not None:
        data["units_flottes"] = units_flottes_kv

    # "Autres revenus" derived algebraically (profit_departemental -
    # profit_brut + total_depenses) rather than read off one fixed row --
    # the row that plays this role (a "transfert"/"prorata" adjustment
    # between the pre- and post-transfer department profit lines) carries a
    # different label per department ("TRANSFERT REVENU F&A ET PLAN DE
    # PROTECTION" for Véhicules neufs vs "TRANSFERT PROFIT BRUT PIÈCES AU
    # CLIENT" for Pièces, etc.), so deriving it from figures already read
    # above holds regardless of the exact wording. Verified against all 3
    # HAWKS files: identity holds to the dollar for every department/column.
    pb = data.get("profit_brut", {}).get("real")
    td = data.get("total_depenses", {}).get("real")
    pd = data.get("profit_departemental", {}).get("real")
    if isinstance(pb, (int, float)) and isinstance(td, (int, float)) and isinstance(pd, (int, float)):
        data["autres_revenus"] = real_only_kv(pd - pb + td)

    data["line_items"] = hawks_line_items(ws, money_col)
    return data

def hawks_summary_section(ws2, col):
    return {
        "ventes_nettes": real_only_kv(ws2.cell(row=4, column=col).value),
        "profit_brut": real_only_kv(ws2.cell(row=5, column=col).value),
        "total_depenses": real_only_kv(ws2.cell(row=60, column=col).value),
        "total_autres_revenus": real_only_kv(ws2.cell(row=62, column=col).value),
        "profit_net": real_only_kv(ws2.cell(row=66, column=col).value),
        "impot": real_only_kv(ws2.cell(row=67, column=col).value),
        "profit_net_apres_impot": real_only_kv(ws2.cell(row=68, column=col).value),
    }

def extract_hawks_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    tag = wb["Page2"]["U1"].value if "Page2" in wb.sheetnames else None
    m = re.search(r'(\d{4})(\d{2})\s*-\s*PAGE', str(tag) or '')
    if not m:
        raise ValueError(f"Could not find the '<code> - YYYYMM - PAGE N' period tag in {path}")
    year, month_num = int(m.group(1)), int(m.group(2))
    month_name = HAWKS_MONTHS[month_num]
    company = wb["Page1"]["J6"].value

    new_units_month, new_units_ytd = hawks_new_vehicle_units(wb)
    fleet_month, fleet_ytd = hawks_fleet_units(wb)
    used_month, used_ytd = hawks_used_vehicle_units(wb)

    # (sheet, month_col, ytd_col, (units_month, units_ytd), (flottes_month, flottes_ytd))
    dept_specs = {
        "Véhicules neufs": ("Page3", 4, 6, (new_units_month, new_units_ytd), (fleet_month, fleet_ytd)),
        "Véhicules usagés": ("Page3", 8, 10, (used_month, used_ytd), (None, None)),
        "Service": ("Page4", 4, 6, (None, None), (None, None)),
        "Carrosserie": ("Page4", 8, 10, (None, None), (None, None)),
        "Pièces": ("Page4", 12, 14, (None, None), (None, None)),
    }
    departments_month, departments_ytd = {}, {}
    for name, (sheet, mcol, ycol, (u_m, u_y), (f_m, f_y)) in dept_specs.items():
        departments_month[name] = hawks_department_section(wb, sheet, mcol, units_kv=u_m, units_flottes_kv=f_m)
        departments_ytd[name] = hawks_department_section(wb, sheet, ycol, units_kv=u_y, units_flottes_kv=f_y)

    ws2 = wb["Page2"]
    summary_month = hawks_summary_section(ws2, 7)   # G = MOIS
    summary_ytd = hawks_summary_section(ws2, 10)    # J = CUMUL ANNUEL

    kpis_month = {}
    kpis_month.update(department_kpis_for_company_view(departments_month))
    kpis_month.update(summary_kpis_for_company_view(summary_month))
    kpis_month.update(expense_breakdown_kpis_for_company_view(departments_month))
    kpis_month.update(unit_economics_kpis_for_company_view(kpis_month))
    kpis_ytd = {}
    kpis_ytd.update(department_kpis_for_company_view(departments_ytd))
    kpis_ytd.update(summary_kpis_for_company_view(summary_ytd))
    kpis_ytd.update(expense_breakdown_kpis_for_company_view(departments_ytd))
    kpis_ytd.update(unit_economics_kpis_for_company_view(kpis_ytd))

    return {
        "company": company.strip() if isinstance(company, str) else company,
        "month_name": month_name,
        "month_num": month_num,
        "year": year,
        "period_key": f"{year}-{month_num:02d}",
        "sections": {
            "month": {"source_title": month_name, "kpis": kpis_month, "departments": departments_month},
            "ytd": {"source_title": f"AAD {month_name}", "kpis": kpis_ytd, "departments": departments_ytd},
        },
    }

def extract_any_file(path):
    """HAWKS sends GM's standardized .xlsm composite statement; every other
    dealer sends the Quotus-based .xlsx template. Dispatch on extension."""
    if path.lower().endswith(".xlsm"):
        return extract_hawks_file(path)
    return extract_file(path)

def load_store(store_path):
    if os.path.exists(store_path):
        with open(store_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"dealers": {}}

def save_store(store, store_path):
    # A fresh checkout (e.g. a GitHub Actions runner) won't have data/ at all
    # -- git doesn't track empty directories -- so create it on demand rather
    # than assuming it already exists like a long-lived local workspace does.
    os.makedirs(os.path.dirname(store_path) or ".", exist_ok=True)
    with open(store_path, 'w', encoding='utf-8') as f:
        json.dump(store, f, ensure_ascii=False, indent=2)

def merge_into_store(store, extracted):
    dealer_key, display_name = canonicalize_dealer(extracted["company"])
    period = extracted["period_key"]
    d = store["dealers"].setdefault(dealer_key, {"display_name": display_name, "legal_name": extracted["company"], "periods": {}})
    d["display_name"] = display_name
    d["legal_name"] = extracted["company"]
    d["periods"][period] = {
        "month_name": extracted["month_name"],
        "month_num": extracted["month_num"],
        "year": extracted["year"],
        "sections": extracted["sections"]
    }
    return store

if __name__ == "__main__":
    # Relative to this script's own location (repo_root/src/extract.py ->
    # repo_root/data/data.json) so this runs identically whether invoked
    # locally or from a GitHub Actions checkout at a different path.
    repo_root = Path(__file__).resolve().parent.parent
    store_path = str(repo_root / "data" / "data.json")
    store = load_store(store_path)
    skipped = []
    for path in sys.argv[1:]:
        # sources/ est maintenant alimenté automatiquement par drive_sync.py,
        # qui récupère TOUT ce qui se trouve dans les dossiers Drive de
        # chaque concessionnaire -- pas seulement les rapports "Réalisé"
        # mensuels que ce script sait lire. Des fichiers annexes (2 pagers,
        # analyses de gross, bonis cadres, inventaires, rapports de ventes
        # privées, etc.) s'y retrouvent donc aussi. On ignore ces fichiers
        # avec un avertissement plutôt que de faire échouer toute la
        # reconstruction du tableau de bord à cause d'un seul fichier
        # inattendu.
        try:
            extracted = extract_any_file(path)
        except Exception as exc:  # noqa: BLE001 - on veut logguer et continuer
            skipped.append((path, str(exc)))
            print(f"IGNORÉ (format non reconnu): {path} -- {exc}", file=sys.stderr)
            continue
        print(f"Extracted: {extracted['company']} - {extracted['period_key']} -- sections: {list(extracted['sections'].keys())}")
        merge_into_store(store, extracted)
    save_store(store, store_path)
    print("Saved store to", store_path)
    if skipped:
        print(f"--- {len(skipped)} fichier(s) ignoré(s) (format non reconnu) ---", file=sys.stderr)
        for path, reason in skipped:
            print(f"  - {path}: {reason}", file=sys.stderr)
